import urllib.request, json, csv, io, sys, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

resp = urllib.request.urlopen('http://localhost:8000/api/authors/filters', timeout=10)
big_schools = sorted(json.loads(resp.read())['schools'])
resp2 = urllib.request.urlopen('http://localhost:8000/api/authors', timeout=10)
authors = json.loads(resp2.read())['authors']

author_era = {}
for a in authors:
    era = a.get('era','')
    if era:
        is_bc = '公元前' in era or '前' in era
        m = re.search(r'(\d{3,4})', era)
        if m:
            y = int(m.group(1))
            author_era[a['name']] = -y if is_bc else y
        else:
            author_era[a['name']] = 9999
    else:
        author_era[a['name']] = 9999

normMap = {
    '存在主义先驱':'存在主义','存在哲学':'存在主义','文学哲学':'存在主义',
    '柏拉图主义':'古希腊哲学','逍遥学派':'古希腊哲学','伊壁鸠鲁主义':'古希腊哲学',
    '米利都学派':'古希腊哲学','埃利亚派':'古希腊哲学','前苏格拉底':'古希腊哲学',
    '古代哲学':'古希腊哲学','犬儒学派':'古希腊哲学','自然哲学':'古希腊哲学',
    '新柏拉图主义':'古希腊哲学','折衷主义':'古希腊哲学','元素论':'古希腊哲学',
    '斯多葛派':'斯多葛学派','斯多葛主义':'斯多葛学派','晚期斯多亚':'斯多葛学派',
    '批判哲学':'德国古典哲学','德国唯心论':'德国古典哲学','唯意志论':'德国古典哲学',
    '交往理论':'法兰克福学派','文化批评':'法兰克福学派',
    '结构马克思主义':'马克思主义','政治经济学':'政治哲学','宗教社会学':'社会学',
    '现实主义政治哲学':'政治哲学','文艺复兴人文主义':'启蒙运动','逻辑实证主义':'实证主义',
    '启蒙哲学':'启蒙运动','启蒙思想':'启蒙运动','苏格兰启蒙':'启蒙运动','人文主义':'启蒙运动',
    '精神分析':'精神分析学','分析心理学':'精神分析学','心理治疗':'精神分析学',
    '逻辑原子主义':'分析哲学','逻辑实用主义':'分析哲学','逻辑实证':'分析哲学',
    '日常语言':'分析哲学','语言哲学':'分析哲学',
    '形式社会学':'社会学','社会心理学':'社会学','群体心理学':'社会学','社会达尔文':'社会学',
    '激进平等':'政治哲学','责任伦理':'政治哲学','社会契约论':'政治哲学','古典经济学':'政治哲学',
    '德性伦理':'伦理学','批判理性主义':'科学哲学',
    '解释学':'现象学','身体哲学':'现象学','意向性':'现象学',
    '常识实在论':'实在论','人本唯物论':'实在论','机械唯物主义':'实在论',
    '结构语言学':'结构主义','进步教育':'实用主义','新实用主义':'实用主义',
    '荒诞文学':'荒诞哲学','浪漫主义先驱':'浪漫主义',
    '近代哲学之父':'近代哲学','有机体哲学':'过程哲学',
    '后现代哲学':'后现代主义','解构主义':'后现代主义',
    '绝对唯心主义':'唯心主义','历史唯物主义':'马克思主义',
    '法兰克福学派（批判理论）':'法兰克福学派','悲观主义哲学':'德国古典哲学','文化霸权理论':'西方马克思主义',
}
def normalize(s): return normMap.get(s.strip(), s.strip())

chrono_order = {
    '古希腊哲学':1,'怀疑论':2,'斯多葛学派':3,'教父哲学':4,'经院哲学':5,'托马斯主义':6,'唯名论':7,
    '近代哲学':10,'理性主义':11,'经验主义':12,'实在论':13,'启蒙运动':14,'浪漫主义':15,
    '德国古典哲学':16,'唯心主义':17,'自由主义':18,'功利主义':19,
    '超验主义':20,'实证主义':21,'社会学':22,
    '马克思主义':23,'生命哲学':24,'实用主义':25,'精神分析学':26,
    '现象学':27,'存在主义':28,'分析哲学':29,'过程哲学':30,'哲学人类学':31,
    '西方马克思主义':32,'法兰克福学派':33,'批判理论':34,
    '科学哲学':35,'荒诞哲学':36,'基督教哲学':37,'宗教哲学':38,'伦理学':39,
    '结构主义':40,'政治哲学':41,'哲学诠释学':42,
    '后结构主义':43,'后现代主义':44,'女性主义':45,'社群主义':46,'技术哲学':47,
}

sub_chrono = {}
for i, k in enumerate([
    '前苏格拉底','米利都学派','埃利亚派','自然哲学','元素论','古希腊哲学',
    '柏拉图主义','逍遥学派','伊壁鸠鲁主义','犬儒学派','斯多葛派','斯多葛主义',
    '折衷主义','古代哲学','新柏拉图主义','晚期斯多亚',
    '教父哲学','经院哲学','托马斯主义','唯名论',
    '近代哲学之父','理性主义','经验主义','常识实在论','机械唯物主义',
    '启蒙哲学','启蒙思想','苏格兰启蒙','人文主义','浪漫主义先驱',
    '批判哲学','德国唯心论','绝对唯心主义','唯意志论',
    '功利主义','自由主义','古典经济学','社会达尔文',
    '实证主义','逻辑实证主义','逻辑实证','逻辑原子主义','逻辑实用主义',
    '马克思主义','历史唯物主义','结构马克思主义','文化霸权理论',
    '实用主义','进步教育','新实用主义',
    '精神分析','分析心理学','心理治疗',
    '现象学','解释学','身体哲学','意向性',
    '存在主义','存在哲学','文学哲学','荒诞文学',
    '分析哲学','日常语言','语言哲学',
    '社会学','形式社会学','社会心理学','群体心理学',
    '西方马克思主义','批判理论','法兰克福学派','交往理论','文化批评',
    '科学哲学','批判理性主义',
    '结构主义','结构语言学',
    '政治哲学','现实主义政治哲学','激进平等','社会契约论',
    '后结构主义','解构主义','后现代哲学','后现代主义',
    '过程哲学','超验主义','生命哲学','实在论',
    '德性伦理','伦理学','政治经济学','责任伦理',
    '技术哲学','女性主义','社群主义',
]): sub_chrono[k] = i+1

tree = {}
for a in authors:
    raw = a.get('school','')
    if not raw: continue
    for s in raw.replace('、','/').replace('，','/').replace(',','/').split('/'):
        s = s.strip()
        if not s: continue
        big = normalize(s)
        if big not in tree: tree[big] = {}
        if s not in tree[big]: tree[big][s] = []
        if a['name'] not in tree[big][s]:
            tree[big][s].append(a['name'])

sorted_bigs = sorted(big_schools, key=lambda s: chrono_order.get(s, 99))

rows = []
for big in sorted_bigs:
    if big in tree:
        subs = sorted(tree[big].items(), key=lambda x: sub_chrono.get(x[0], 500))
        for sub, auths in subs:
            auths_sorted = sorted(auths, key=lambda n: author_era.get(n, 9999))
            rows.append([big, sub] + auths_sorted)

max_cols = max(len(r) for r in rows) if rows else 3

# Write Excel with merged cells
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
wb = Workbook()
ws = wb.active
ws.title = '流派-作者'

# Header
headers = ['大流派','小流派'] + [f'作家{i+1}' for i in range(max_cols-2)]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# Data rows
for i, r in enumerate(rows, 2):
    for j, v in enumerate(r):
        ws.cell(row=i, column=j+1, value=v)

# Merge big school cells
current_big = None
merge_start = 2
for i, r in enumerate(rows):
    big = r[0]
    if big != current_big:
        if current_big and merge_start < i+1:
            ws.merge_cells(start_row=merge_start, start_column=1, end_row=i+1, end_column=1)
        current_big = big
        merge_start = i+2
# Last group
if current_big and merge_start <= len(rows)+1:
    ws.merge_cells(start_row=merge_start, start_column=1, end_row=len(rows)+1, end_column=1)

# Center align
for row in ws.iter_rows(min_row=2, max_row=len(rows)+1, min_col=1, max_col=1):
    for cell in row:
        cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

path = r'C:\Users\wqx_0\PyCharmProjects\Q&ASystem\schools_authors.xlsx'
wb.save(path)
print(f'Saved {len(rows)} rows to {path}')
